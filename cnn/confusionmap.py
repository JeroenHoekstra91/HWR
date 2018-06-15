from keras.preprocessing.image import img_to_array
from keras.models import load_model
from util.character_map import char_map
import cv2
import numpy as np
import argparse
import os
from openpyxl import *
from openpyxl.styles import *
import colorsys

confusionmap = np.zeros((27, 27), dtype = np.uint8)

ap = argparse.ArgumentParser()
ap.add_argument("-d", "--dataset", required=True,
	help="path to input dataset")
ap.add_argument("-m", "--model", required=True,
	help="path to output model")
args = vars(ap.parse_args())

imagePaths = []
for root, dirs, files in os.walk(args["dataset"]):
    for name in files:
        imagePaths.append(os.path.join(root, name))

labels = []
for imagePath in imagePaths:
    label = imagePath.split(os.path.sep)[-2]
    labels.append(label)

# load the trained convolutional neural network
print("[INFO] loading network...")
model = load_model(args["model"])

loss = 0.0
for i in range(len(imagePaths)):
	# load the image
	image = cv2.imread(imagePaths[i])

	# pre-process the image for classification
	image = cv2.resize(image, (28, 28))
	image = image.astype("float") / 255.0
	image = img_to_array(image)
	image = np.expand_dims(image, axis=0)

	# classify the input image
	prediction = list(model.predict(image)[0])
	argmax = np.argmax(prediction)
	predicted_label = char_map.keys()[char_map.values().index(argmax)]
	confusionmap[char_map[labels[i]], argmax] += 1

	if predicted_label != labels[i]:
		loss = loss + 1
		print("Should be: %s -> Predicted: %s" % (labels[i], predicted_label))

print("\nConfusion Map:")
print(confusionmap)

print("\nSample size: %i" % (len(imagePaths)))
print("Total loss: %i" % (loss))
print("Accuracy: " + str(100 - loss/len(imagePaths)*100) + "%")
print("Loss:" + str(loss/len(imagePaths)*100) + "%")

# Write confusion map to excel sheet
confusion_color = "0011FF"

workbook = Workbook()
sheet = workbook.active
sheet.title = "Confusion Map"
hsv = colorsys.rgb_to_hsv(
    int(confusion_color[0:2], 16)/255.0,
    int(confusion_color[2:4], 16)/255.0,
    int(confusion_color[4:6], 16)/255.0)

for i in range(27):
    # Write labels
    cell = sheet.cell(column=i+2, row=1, value=sorted(char_map.keys())[i])
    cell.alignment=Alignment(text_rotation=90, horizontal='center')
    sheet.cell(column=1, row=i+2, value=sorted(char_map.keys())[i])

    min_confusion = confusionmap[i][0:27].min()
    max_confusion = confusionmap[i][0:27].max()
    # write cell values
    for j in range(27):
        col = j + 2
        row = i + 2
        value = confusionmap[i][j]
        confusion = value / (max_confusion - min_confusion * 1.0) 
        
        rgb = colorsys.hsv_to_rgb(hsv[0], confusion, hsv[2])
        color = str(format(int(rgb[0]*255), '02X')) + str(format(int(rgb[1]*255), '02X')) + str(format(int(rgb[2]*255), '02X'))
        
        cell = sheet.cell(column=col, row=row, value=value)
        cell.fill = PatternFill(fgColor=color, fill_type="solid")

sheet.cell(column=1, row=1, value="real/predicted")

sheet.cell(column=1, row=30, value="Sample size:").font = Font(bold=True)
sheet.cell(column=3, row=30, value=str(len(imagePaths))).font = Font(bold=True)

sheet.cell(column=1, row=31, value="Total loss:").font = Font(bold=True)
sheet.cell(column=3, row=31, value=str(int(loss))).font = Font(bold=True)

sheet.cell(column=1, row=32, value="Accuracy:").font = Font(bold=True)
sheet.cell(column=3, row=32, value=str(100 - loss/len(imagePaths)*100) + "%").font = Font(bold=True)

sheet.cell(column=1, row=33, value="Loss:").font = Font(bold=True)
sheet.cell(column=3, row=33, value=str(loss/len(imagePaths)*100) + "%").font = Font(bold=True)

workbook.save(filename = "confusionmap.xlsx")